from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

wb = Workbook()

# ---------- Siparişler sayfası ----------
ws = wb.active
ws.title = "Siparişler"

# Sütun başlıkları (kullanıcının verdiği sıra)
headers = [
    "İŞ SON DURUMU", "TARİH", "SİPARİŞ ALINAN FİRMA", "ÜRÜN ADI", "RENK",
    "ADET", "BİRİM FİYAT (TL)", "ÇANTA FİYATI (TL)", "VERİLEN FİYAT",
    "KALAN PARA", "ALINAN ÖDEME", "KALAN ÖDEME"
]
ws.append(headers)

# Stil: başlık
header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(bold=True, color="FFFFFF", size=11)
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for c in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=c)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

# Örnek veri (5 satır)
sample = [
    ["Tamamlandı", date(2026, 8, 1), "ABC Tekstil", "Dikdörtgen Çanta", "Kırmızı", 100, 25.0, 5.0, 3000.0, None, 3000.0, None],
    ["Devam Ediyor", date(2026, 8, 5), "XYZ Moda", "Yuvarlak Çanta", "Mavi", 50, 30.0, 6.0, 1500.0, None, 1000.0, None],
    ["Beklemede", date(2026, 8, 10), "Delta Sanayi", "Kare Çanta", "Yeşil", 200, 20.0, 4.0, 4000.0, None, 2000.0, None],
    ["Tamamlandı", date(2026, 8, 12), "ABC Tekstil", "Üçgen Çanta", "Siyah", 80, 28.0, 5.5, 2240.0, None, 2240.0, None],
    ["Devam Ediyor", date(2026, 8, 15), "Nova A.Ş.", "Oval Çanta", "Beyaz", 120, 22.0, 4.5, 2640.0, None, 1500.0, None],
]

# Formül sütunları:
# KALAN PARA (J=10) = VERİLEN FİYAT (I=9) - (BİRİM FİYAT (G=7)*ADET (F=6) + ÇANTA FİYATI (H=8)*ADET (F=6))
# KALAN ÖDEME (L=12) = VERİLEN FİYAT (I=9) - ALINAN ÖDEME (K=11)
for r in sample:
    ws.append(r)
    row = ws.max_row
    # TARİH formatı
    ws.cell(row=row, column=2).number_format = "DD.MM.YYYY"
    # para formatları
    for col in (7, 8, 9, 10, 11, 12):
        ws.cell(row=row, column=col).number_format = "#,##0.00 ₺"
    # KALAN PARA formülü
    ws.cell(row=row, column=10).value = f"=I{row}-(G{row}*F{row}+H{row}*F{row})"
    # KALAN ÖDEME formülü
    ws.cell(row=row, column=12).value = f"=I{row}-K{row}"
    for c in range(1, len(headers) + 1):
        ws.cell(row=row, column=c).border = border

# Sütun genişlikleri
widths = [14, 12, 22, 16, 10, 8, 16, 16, 14, 14, 14, 14]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ---------- Özet sayfası ----------
ws2 = wb.create_sheet("Özet")
ws2.append(["Sipariş Takip Özeti"])
ws2["A1"].font = Font(bold=True, size=14, color="1F4E78")
ws2.append([])
ws2.append(["Gösterge", "Değer"])
for c in (1, 2):
    ws2.cell(row=3, column=c).font = Font(bold=True, color="FFFFFF")
    ws2.cell(row=3, column=c).fill = header_fill

data_rows = [
    ("Toplam Sipariş", "=COUNTA(Siparişler!A2:A100000)"),
    ("Toplam Verilen Fiyat (TL)", "=SUM(Siparişler!I2:I100000)"),
    ("Toplam Alınan Ödeme (TL)", "=SUM(Siparişler!K2:K100000)"),
    ("Toplam Kalan Ödeme (TL)", "=SUM(Siparişler!L2:L100000)"),
    ("Toplam Kalan Para (TL)", "=SUM(Siparişler!J2:J100000)"),
    ("Firma Sayısı", "=SUMPRODUCT(1/COUNTIF(Siparişler!C2:C100000,Siparişler!C2:C100000&\"\"))"),
]
for label, formula in data_rows:
    ws2.append([label, formula])
    ws2.cell(row=ws2.max_row, column=2).number_format = "#,##0.00 ₺"

ws2.append([])
ws2.append(["Durum Dağılımı"])
ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True)
ws2.append(["Durum", "Adet"])
for c in (1, 2):
    rr = ws2.max_row
    ws2.cell(row=rr, column=c).font = Font(bold=True, color="FFFFFF")
    ws2.cell(row=rr, column=c).fill = header_fill
for durum in ["Tamamlandı", "Devam Ediyor", "Beklemede"]:
    ws2.append([durum, f'=COUNTIF(Siparişler!A2:A100000,"{durum}")'])

for col, w in [(1, 30), (2, 18)]:
    ws2.column_dimensions[get_column_letter(col)].width = w

# ---------- Kaydet ----------
out = "C:/Users/ASUS/Desktop/Projeler/office-canvas-tasks/excel/siparis-takip.xlsx"
wb.save(out)
print("KAYDEDILDI:", out)
print("Sayfalar:", wb.sheetnames)
