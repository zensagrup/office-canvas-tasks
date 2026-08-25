"""
Sipariş Takip Web Uygulaması — FastAPI + openpyxl
Veri G:\\Drive'ım\\SİPARİŞ TAKİP\\siparis.xlsx içinde saklanır.
"""
import os
from fastapi import FastAPI, HTTPException, Request
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.responses import HTMLResponse, JSONResponse
from pydantic import BaseModel
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import shutil

BASE = r"G:\Drive'ım\SİPARİŞ TAKİP"
EXCEL = os.path.join(BASE, "siparis.xlsx")
TEMPLATES = os.path.join(BASE, "templates")
STATIC = os.path.join(BASE, "static")

os.makedirs(TEMPLATES, exist_ok=True)
os.makedirs(STATIC, exist_ok=True)

HEADERS = [
    "ID", "İŞ SON DURUMU", "TARİH", "SİPARİŞ ALINAN FİRMA", "ÜRÜN ADI", "RENK",
    "ADET", "BİRİM FİYAT (TL)", "ÇANTA FİYATI (TL)", "VERİLEN FİYAT",
    "KALAN PARA", "ALINAN ÖDEME", "KALAN ÖDEME", "OLUŞTURMA"
]

app = FastAPI(title="Sipariş Takip")
templates = Jinja2Templates(directory=TEMPLATES)
app.mount("/static", StaticFiles(directory=STATIC), name="static")

# ---------- Excel yardımcıları ----------
def init_excel():
    if os.path.exists(EXCEL):
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Siparişler"
    ws.append(HEADERS)
    fill = PatternFill("solid", fgColor="1F4E78")
    fnt = Font(bold=True, color="FFFFFF")
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill; cell.font = fnt
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    widths = [5,14,12,22,16,10,8,16,16,14,14,14,14,18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    wb.save(EXCEL)

def load_ws():
    init_excel()
    return load_workbook(EXCEL)

def next_id(ws):
    maxid = 0
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, int) and v > maxid:
            maxid = v
    return maxid + 1

def calc_row(d, row_id):
    """KALAN PARA ve KALAN ÖDEME hesapla (değer olarak)."""
    birim = float(d.get("birim_fiyat") or 0)
    cantaf = float(d.get("cantafi_fiyat") or 0)
    adet = float(d.get("adet") or 0)
    verilen = float(d.get("verilen_fiyat") or 0)
    alinan = float(d.get("alinan_odeme") or 0)
    kalan_para = verilen - (birim * adet + cantaf * adet)
    kalan_odeme = verilen - alinan
    return kalan_para, kalan_odeme

# ---------- Pydantic model ----------
class SiparisIn(BaseModel):
    is_son_durumu: str = ""
    tarih: str = ""           # DD.MM.YYYY veya ISO
    firma: str = ""
    urun_adi: str = ""
    renk: str = ""
    adet: float = 0
    birim_fiyat: float = 0
    cantafi_fiyat: float = 0
    verilen_fiyat: float = 0
    alinan_odeme: float = 0

# ---------- Routes ----------
@app.get("/", response_class=HTMLResponse)
def index(request: Request):
    return templates.TemplateResponse(request=request, name="index.html")

@app.get("/api/siparisler")
def list_siparisler():
    wb = load_ws()
    ws = wb["Siparişler"]
    rows = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, len(HEADERS) + 1)]
        if vals[0] is None:
            continue
        rows.append(dict(zip(
            ["id","is_son_durumu","tarih","firma","urun_adi","renk","adet",
             "birim_fiyat","cantafi_fiyat","verilen_fiyat","kalan_para",
             "alinan_odeme","kalan_odeme","olusturma"], vals)))
    return rows

@app.post("/api/siparisler")
def create_siparis(s: SiparisIn):
    wb = load_ws()
    ws = wb["Siparişler"]
    rid = next_id(ws)
    kalan_para, kalan_odeme = calc_row(s.model_dump(), rid)
    tarih_fmt = s.tarih if s.tarih else datetime.now().strftime("%d.%m.%Y")
    row = [
        rid, s.is_son_durumu, tarih_fmt, s.firma, s.urun_adi, s.renk,
        s.adet, s.birim_fiyat, s.cantafi_fiyat, s.verilen_fiyat,
        round(kalan_para, 2), s.alinan_odeme, round(kalan_odeme, 2),
        datetime.now().strftime("%d.%m.%Y %H:%M")
    ]
    ws.append(row)
    wb.save(EXCEL)
    return {"id": rid, "kalan_para": kalan_para, "kalan_odeme": kalan_odeme}

@app.put("/api/siparisler/{sid}")
def update_siparis(sid: int, s: SiparisIn):
    wb = load_ws()
    ws = wb["Siparişler"]
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == sid:
            kalan_para, kalan_odeme = calc_row(s.model_dump(), sid)
            tarih_fmt = s.tarih if s.tarih else ws.cell(row=r, column=3).value
            vals = [sid, s.is_son_durumu, tarih_fmt, s.firma, s.urun_adi, s.renk,
                    s.adet, s.birim_fiyat, s.cantafi_fiyat, s.verilen_fiyat,
                    round(kalan_para,2), s.alinan_odeme, round(kalan_odeme,2),
                    ws.cell(row=r, column=14).value]
            for i, v in enumerate(vals, 1):
                ws.cell(row=r, column=i).value = v
            wb.save(EXCEL)
            return {"id": sid, "updated": True}
    raise HTTPException(404, "Sipariş bulunamadı")

@app.delete("/api/siparisler/{sid}")
def delete_siparis(sid: int):
    wb = load_ws()
    ws = wb["Siparişler"]
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == sid:
            ws.delete_rows(r, 1)
            wb.save(EXCEL)
            return {"id": sid, "deleted": True}
    raise HTTPException(404, "Sipariş bulunamadı")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8000, reload=False)
